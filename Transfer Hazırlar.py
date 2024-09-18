#Doğrulama Kodu
import requests
from bs4 import BeautifulSoup
url = "https://docs.google.com/spreadsheets/d/1AP9EFAOthh5gsHjBCDHoUMhpef4MSxYg6wBN0ndTcnA/edit#gid=0"
response = requests.get(url)
html_content = response.content
soup = BeautifulSoup(html_content, "html.parser")
first_cell = soup.find("td", {"class": "s2"}).text.strip()
if first_cell != "Aktif":
    exit()
first_cell = soup.find("td", {"class": "s1"}).text.strip()
print(first_cell)


import pandas as pd
import os
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import requests




url = "https://haydigiy.online/Products/rafkodlari.php"
response = requests.get(url)
soup = BeautifulSoup(response.text, "html.parser")
table = soup.find("table")
data = []
for row in table.find_all("tr"):
    row_data = []
    for cell in row.find_all(["th", "td"]):
        row_data.append(cell.get_text(strip=True))
    data.append(row_data)
df = pd.DataFrame(data[1:], columns=data[0])
df.to_excel("Raf Kodları.xlsx", index=False)




url = "https://task.haydigiy.com/FaprikaOrderXls/Q4DKPP/1/"
try:
    response = requests.get(url)
    response.raise_for_status()
    excel_file = "Site Veri.xlsx"
    with open(excel_file, "wb") as f:
        f.write(response.content)
    df = pd.read_excel(excel_file)
    columns_to_keep = ["Id", "UrunAdi", "Barkod", "Adet", "Varyant"]
    df = df[columns_to_keep]
    new_excel_file = "Site Veri 2.xlsx"
    df.to_excel(new_excel_file, index=False)
except requests.exceptions.RequestException as e:
    print("İstek sırasında bir hata oluştu:", e)



duzenlenmis_excel_file = "Site Veri 2.xlsx"
duzenlenmis_df = pd.read_excel(duzenlenmis_excel_file)
google_indirilen_excel_file = "Raf Kodları.xlsx"
google_df = pd.read_excel(google_indirilen_excel_file)

karşılıklar = []
for urun_adi in duzenlenmis_df["UrunAdi"]:
    match = google_df[google_df.iloc[:, 1].str.strip() == urun_adi]
    if not match.empty:
        index = match.index[0]
        karşılıklar.append(google_df.iloc[index, 2])
    else:
        karşılıklar.append("")
duzenlenmis_df["E"] = karşılıklar

son_excel_file = "son_dosya.xlsx"
duzenlenmis_df.to_excel(son_excel_file, index=False)

duzenlenmis_excel_file = "son_dosya.xlsx"
duzenlenmis_df = pd.read_excel(duzenlenmis_excel_file)

for index, row in duzenlenmis_df.iterrows():
    if pd.notna(row["UrunAdi"]) and pd.isna(row["E"]):
        duzenlenmis_df.at[index, "E"] = "Raf Kodu Yok"
duzenlenmis_df.to_excel(duzenlenmis_excel_file, index=False)

duzenlenmis_excel_file = "son_dosya.xlsx"
duzenlenmis_df = pd.read_excel(duzenlenmis_excel_file)
grouped = duzenlenmis_df.groupby("Id")
new_dfs = []
for group_name, group_df in grouped:
    temp_df = pd.DataFrame(data=group_df)
    for index, row in temp_df.iterrows():
        if pd.notna(row["UrunAdi"]) and pd.notna(row["E"]):
            matching_row = duzenlenmis_df[(duzenlenmis_df["UrunAdi"] == row["UrunAdi"]) & (duzenlenmis_df["E"] == row["E"])]
            if not matching_row.empty:
                temp_df.at[index, "E"] = matching_row.iloc[0]["E"]
    new_excel_file = f"{group_name}.xlsx"
    temp_df.to_excel(new_excel_file, index=False)
    new_dfs.append(temp_df)
for idx, df in enumerate(new_dfs):
    new_excel_file = f"{df['Id'].iloc[0]}.xlsx"
    df.to_excel(new_excel_file, index=False)
for filename in os.listdir():
    if filename.endswith(".xlsx") and not any(str(df['Id'].iloc[0]) in filename for df in new_dfs):
        os.remove(filename)
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        e_column = "E"
        if e_column in df.columns:
            df[e_column + "_copy"] = df[e_column]
            df.to_excel(excel_file, index=False)
        else:
            print(f"{excel_file} dosyası {e_column} sütunu eksik, işlem atlandı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        e_copy_column = "E_copy"
        if e_copy_column in df.columns:
            df[e_copy_column] = df[e_copy_column].str.split("-").str[0].str.strip()
            df[e_copy_column] = pd.to_numeric(df[e_copy_column], errors="coerce")
            df = df.sort_values(by=[e_copy_column])
            df.to_excel(excel_file, index=False)
        else:
            print(f"{excel_file} dosyası {e_copy_column} sütunu eksik, işlem atlandı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        e_copy_column = "E_copy"
        if e_copy_column in df.columns:
            df = df.drop(columns=[e_copy_column])
            df.to_excel(excel_file, index=False)
        else:
            print(f"{excel_file} dosyası {e_copy_column} sütunu eksik, işlem atlandı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        variant_column = "Varyant"
        if variant_column in df.columns:
            df[variant_column] = df[variant_column].str.replace("Beden:", "").str.strip()
            df.to_excel(excel_file, index=False)
        else:
            print(f"{excel_file} dosyası {variant_column} sütunu eksik, işlem atlandı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        id_column = "Id"
        varyant_column = "Varyant"
        if id_column in df.columns and varyant_column in df.columns:
            df["UrunAdi"] = df["UrunAdi"] + " //" + df[varyant_column]
            df = df.drop(columns=[id_column, varyant_column])
            df.to_excel(excel_file, index=False)
        else:
            print(f"{excel_file} dosyası {id_column} veya {varyant_column} sütunu eksik, işlem atlandı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        e_column = "E"
        if e_column in df.columns:
            e_values = df[e_column]
            df = df.drop(columns=[e_column])
            df.insert(0, e_column, e_values)
            df.to_excel(excel_file, index=False)
        else:
            print(f"{excel_file} dosyası {e_column} sütunu eksik, işlem atlandı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        df = pd.read_excel(excel_file)
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i, column in enumerate(df.columns):
                column_width = max(df[column].astype(str).apply(len).max(), len(column)) + 5
                worksheet.set_column(i, i, column_width)
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        wb = load_workbook(excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(size=14)
        wb.save(excel_file)
    except Exception as e:
        print(f"Hata oluştu: {e}")
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        wb = load_workbook(excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(size=14, bold=True)
        wb.save(excel_file)
    except Exception as e:
        print(f"Hata oluştu: {e}")
border = Border(left=Side(style="thin"), 
                right=Side(style="thin"), 
                top=Side(style="thin"), 
                bottom=Side(style="thin"))
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        wb = load_workbook(excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(size=14, bold=True)
                    cell.border = border
        wb.save(excel_file)
    except Exception as e:
        print(f"Hata oluştu: {e}")
border = Border(left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"))
table_style = TableStyleInfo(
    name="TableStyleLight1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
excel_files = [filename for filename in os.listdir() if filename.endswith(".xlsx")]
for excel_file in excel_files:
    try:
        wb = load_workbook(excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws[1]:
                if cell.value == "E":
                    cell.value = "Raf Kodu"
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(size=14, bold=True)
                    cell.border = border
        table = Table(displayName="MyTable", ref=ws.dimensions, tableStyleInfo=table_style)
        ws.add_table(table)
        wb.save(excel_file)
    except Exception as e:
        print(f"Hata oluştu: {e}")